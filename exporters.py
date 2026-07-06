#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export formats module (CSV, Excel, HTML, JSON, Markdown)."""

import json
import csv
from pathlib import Path
from typing import Any, Dict, List

from utils import extract_plain_text, extract_sender_name, extract_timestamp


class BaseExporter:
    """Base exporter class."""

    def __init__(
        self, messages: List[Dict[str, Any]], id_index: Dict[str, Dict[str, Any]]
    ):
        self.messages = [m for m in messages if m.get("type") == "message"]
        self.id_index = id_index

    def _format_message_record(self, message: Dict[str, Any]) -> Dict[str, str]:
        """Convert message to structured record."""
        timestamp = extract_timestamp(message)
        sender = extract_sender_name(message)
        text = extract_plain_text(message.get("text")).strip()
        msg_id = str(message.get("id", ""))
        reply_id = str(message.get("reply_to_message_id", ""))

        return {
            "id": msg_id,
            "timestamp": timestamp,
            "sender": sender,
            "text": text,
            "reply_to_id": reply_id,
        }

    def export(self, output_path: Path) -> None:
        """Export messages. Override in subclasses."""
        raise NotImplementedError


class CSVExporter(BaseExporter):
    """Export to CSV format."""

    def export(self, output_path: Path) -> None:
        """Export messages to CSV."""
        records = [self._format_message_record(m) for m in self.messages]

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f, fieldnames=["id", "timestamp", "sender", "text", "reply_to_id"]
            )
            writer.writeheader()
            writer.writerows(records)


class JSONExporter(BaseExporter):
    """Export to JSON format."""

    def export(self, output_path: Path) -> None:
        """Export messages to JSON."""
        records = [self._format_message_record(m) for m in self.messages]
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)


class MarkdownExporter(BaseExporter):
    """Export to Markdown format."""

    def export(self, output_path: Path) -> None:
        """Export messages to Markdown."""
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("# Telegram Chat Export\n\n")

            for message in self.messages:
                timestamp = extract_timestamp(message)
                sender = extract_sender_name(message)
                text = extract_plain_text(message.get("text")).strip()

                f.write(f"**{sender}** — `{timestamp}`\n\n")

                if text:
                    f.write(f"{text}\n\n")
                else:
                    f.write("[رسانه]\n\n")

                f.write("---\n\n")


class HTMLExporter(BaseExporter):
    """Export to HTML format."""

    def export(self, output_path: Path) -> None:
        """Export messages to HTML."""
        html_lines = [
            "<!DOCTYPE html>",
            '<html dir="rtl" lang="fa">',
            "<head>",
            '<meta charset="UTF-8">',
            '<meta name="viewport" content="width=device-width, initial-scale=1">',
            "<title>Telegram Chat Export</title>",
            "<style>",
            "body { font-family: Arial, sans-serif; background: #f5f5f5; padding: 20px; }",
            ".container { max-width: 800px; margin: 0 auto; }",
            ".message { background: white; padding: 15px; margin: 10px 0; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }",
            ".sender { font-weight: bold; color: #0084ff; margin-bottom: 5px; }",
            ".timestamp { color: #999; font-size: 12px; margin-right: 10px; }",
            ".text { color: #333; line-height: 1.6; }",
            ".media { color: #999; font-style: italic; }",
            "</style>",
            "</head>",
            "<body>",
            '<div class="container">',
            "<h1>Telegram Chat Export</h1>",
        ]

        for message in self.messages:
            timestamp = extract_timestamp(message)
            sender = extract_sender_name(message)
            text = extract_plain_text(message.get("text")).strip()

            html_lines.append('<div class="message">')
            html_lines.append(
                f'<div><span class="sender">{sender}</span><span class="timestamp">{timestamp}</span></div>'
            )

            if text:
                # Escape HTML
                text_safe = (
                    text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                )
                html_lines.append(f'<div class="text">{text_safe}</div>')
            else:
                html_lines.append('<div class="media">[رسانه]</div>')

            html_lines.append("</div>")

        html_lines.extend(["</div>", "</body>", "</html>"])

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(html_lines))


class ExcelExporter(BaseExporter):
    """Export to Excel format (requires openpyxl)."""

    def export(self, output_path: Path) -> None:
        """Export messages to Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            import logging

            logging.warning(
                "⚠️  openpyxl not installed. Install with: pip install openpyxl"
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"

        # Headers
        headers = ["ID", "Date/Time", "Sender", "Message", "Reply To ID"]
        ws.append(headers)

        # Format header row
        header_fill = PatternFill(
            start_color="0084ff", end_color="0084ff", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for message in self.messages:
            record = self._format_message_record(message)
            ws.append(
                [
                    record["id"],
                    record["timestamp"],
                    record["sender"],
                    record["text"][:500],  # Truncate long messages
                    record["reply_to_id"],
                ]
            )

        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 50
        ws.column_dimensions["E"].width = 10

        # Wrap text in message column
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(output_path)


class TXTExporter(BaseExporter):
    """Export to plain text format with reply context."""

    def export(self, output_path: Path) -> None:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("Telegram Chat Export\n")
            f.write("=" * 50 + "\n\n")
            for message in self.messages:
                record = self._format_message_record(message)
                ts = record["timestamp"]
                sender = record["sender"]
                text = record["text"]
                reply_id = record.get("reply_to_id", "")

                f.write(f"[{ts}] {sender}\n")

                if reply_id and reply_id in self.id_index:
                    parent = self.id_index[reply_id]
                    parent_sender = extract_sender_name(parent)
                    parent_text = extract_plain_text(parent.get("text")).strip()
                    if not parent_text:
                        parent_text = "[media]"
                    if len(parent_text) > 80:
                        parent_text = parent_text[:80] + "..."
                    f.write(f"  \u21b3 {parent_sender}: {parent_text}\n")

                f.write(f"{text or '[media]'}\n")
                f.write("-" * 40 + "\n\n")


def get_exporter(format_type: str) -> type:
    """Get exporter class by format name."""
    exporters = {
        "csv": CSVExporter,
        "json": JSONExporter,
        "html": HTMLExporter,
        "md": MarkdownExporter,
        "txt": TXTExporter,
        "excel": ExcelExporter,
        "xlsx": ExcelExporter,
    }
    return exporters.get(format_type.lower(), CSVExporter)
